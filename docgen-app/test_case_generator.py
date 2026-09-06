"""
test_case_generator.py
Generates a Test Case workbook (.xlsx) from the canonical JSON.

Produces ~18-22 test cases covering:
  - Positive / happy path scenarios
  - Data validation (mandatory fields, malformed data, duplicates)
  - Connectivity and authentication failures
  - Error handling and recovery (retry, notifications, logging)
  - Work-type-specific edge cases (pagination for API sources,
    4xx/5xx for API targets, corrupted files for file sources,
    constraint violations for DB targets)

TBD content is highlighted in bright yellow so reviewers can spot
cases that need input from the requirement to be completable.
Matches the BRD/TDD/Flow diagram highlighting convention.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import re


_TBD_RE = re.compile(r'\bTBD\b', re.IGNORECASE)

# ── Cell styling (matches BRD/TDD conventions) ──
HEADER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
TBD_FILL    = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
ALT_FILL    = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
TITLE_FONT  = Font(name='Calibri', size=16, bold=True, color='1F3864')
META_FONT   = Font(name='Calibri', size=10, italic=True, color='666666')
CELL_FONT   = Font(name='Calibri', size=10)
TBD_FONT    = Font(name='Calibri', size=10, bold=True)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# (name, width, alignment)
COLUMNS = [
    ("Test ID",         12, 'center'),
    ("Scenario",        35, 'left'),
    ("Type",            14, 'center'),
    ("Priority",        10, 'center'),
    ("Preconditions",   35, 'left'),
    ("Test Steps",      50, 'left'),
    ("Expected Result", 40, 'left'),
    ("Actual Result",   20, 'left'),
    ("Status",          12, 'center'),
    ("Comments",        25, 'left'),
]


def _v(field, default="TBD"):
    if field is None:
        return default
    if isinstance(field, str):
        return field
    if isinstance(field, dict) and "value" in field:
        return str(field["value"])
    return default


def _has_tbd(text):
    return bool(_TBD_RE.search(str(text)))


def _build_test_cases(canonical):
    """Return the ordered list of test-case dicts for the given canonical."""
    src         = _v(canonical.get("source", {}).get("system"))
    tgt         = _v(canonical.get("target", {}).get("system"))
    src_type    = _v(canonical.get("source", {}).get("interfaceType")).lower()
    tgt_type    = _v(canonical.get("target", {}).get("interfaceType")).lower()
    src_auth    = _v(canonical.get("security", {}).get("sourceAuth"))
    tgt_auth    = _v(canonical.get("security", {}).get("targetAuth"))
    entities    = ", ".join(_v(e) for e in canonical.get("entities", [])) or "records"
    trigger     = _v(canonical.get("integration", {}).get("trigger"))
    schedule    = _v(canonical.get("integration", {}).get("schedule"))
    target_tbl  = _v(canonical.get("target", {}).get("targetTable"))
    mapping     = _v(canonical.get("mappings", {}).get("status"))
    sla         = _v(canonical.get("integration", {}).get("sla"))

    cases = []
    counter = [0]

    def _id():
        counter[0] += 1
        return f"TC-{counter[0]:03d}"

    def add(scenario, type_, priority, preconditions, steps, expected, comments=""):
        cases.append({
            "id": _id(),
            "scenario": scenario,
            "type": type_,
            "priority": priority,
            "preconditions": preconditions,
            "steps": steps,
            "expected": expected,
            "comments": comments,
        })

    # ═══ POSITIVE PATH ═══
    add(
        scenario=f"End-to-end successful run with valid {entities}",
        type_="Positive",
        priority="High",
        preconditions=f"{src} contains valid {entities} records; {tgt} reachable; credentials configured (source: {src_auth}, target: {tgt_auth})",
        steps=(
            f"1. Trigger integration ({trigger} — {schedule})\n"
            f"2. Verify data extraction from {src}\n"
            f"3. Verify transformation per field mapping\n"
            f"4. Verify data loaded to {tgt}\n"
            f"5. Check run log for Completed status"
        ),
        expected="All valid records processed; run status = Completed; record count matches source; no errors logged",
        comments="Happy path — must pass before release",
    )

    add(
        scenario=f"Incremental pull — only new/changed {entities} since last successful run",
        type_="Positive",
        priority="High",
        preconditions=f"Previous successful run recorded; new {entities} added to {src} after that timestamp",
        steps=(
            "1. Note timestamp of last successful run\n"
            f"2. Add new {entities} to {src}\n"
            "3. Trigger integration\n"
            "4. Verify only new/changed records are processed"
        ),
        expected="Only records with modified_timestamp > last_run_timestamp are processed; unchanged records skipped",
        comments="Critical for scheduled batch integrations",
    )

    add(
        scenario="Scheduled trigger fires automatically at configured time",
        type_="Positive",
        priority="Medium",
        preconditions=f"Integration deployed; schedule configured as: {schedule}",
        steps=(
            "1. Wait for scheduled time\n"
            "2. Verify integration triggers automatically\n"
            "3. Check run log for auto-triggered entry"
        ),
        expected=f"Integration runs at {schedule} without manual intervention",
        comments="Verify scheduler configuration",
    )

    # ═══ DATA VALIDATION ═══
    add(
        scenario="Empty source — zero records to process",
        type_="Boundary",
        priority="Medium",
        preconditions=f"No new {entities} records exist in {src} since last successful run",
        steps=(
            "1. Trigger integration\n"
            "2. Verify graceful handling of zero-record scenario\n"
            "3. Check run log"
        ),
        expected="Integration completes with status = Completed; zero records processed; no errors; notification (if configured) states zero-record run",
    )

    add(
        scenario=f"Mandatory field missing in a source {entities} record",
        type_="Negative",
        priority="High",
        preconditions=f"One source record has null/missing value for a mandatory field",
        steps=(
            "1. Trigger integration\n"
            "2. Verify the invalid record fails validation\n"
            "3. Confirm record is rejected, not loaded to target\n"
            "4. Verify error is logged with record identifier"
        ),
        expected="Invalid record rejected; other valid records continue processing; error logged; status = Completed-with-errors (or per configured policy)",
        comments="Row-level error handling",
    )

    add(
        scenario="Malformed data / data type mismatch",
        type_="Negative",
        priority="High",
        preconditions="Source contains a record where a field value cannot be cast to the target column type",
        steps=(
            "1. Trigger integration\n"
            "2. Verify transformation catches the type mismatch\n"
            "3. Check error handling behavior"
        ),
        expected=f"Record rejected with clear error identifying the field and expected type. Mapping status: {mapping}",
    )

    add(
        scenario=f"Duplicate {entities} record — verify upsert / no duplicates in {tgt}",
        type_="Boundary",
        priority="High",
        preconditions=f"{entities} record already exists in {tgt}; identical record also present in source",
        steps=(
            "1. Trigger integration\n"
            f"2. Verify no duplicate row is created in {tgt}\n"
            "3. Verify existing record is updated (if upsert logic applies)"
        ),
        expected="No duplicates created; upsert on natural key works correctly",
        comments="Idempotency check",
    )

    # ═══ CONNECTIVITY / AUTHENTICATION ═══
    add(
        scenario=f"Source ({src}) unreachable / connection failure",
        type_="Negative",
        priority="High",
        preconditions=f"{src} is unreachable (network down or service off)",
        steps=(
            "1. Trigger integration\n"
            "2. Verify connection attempt fails cleanly\n"
            "3. Check retry behavior per configured policy\n"
            "4. Verify error notification fires on final failure"
        ),
        expected="Integration retries per policy; on final failure, status = Error; notification sent to configured recipients",
    )

    add(
        scenario=f"Target ({tgt}) unreachable during load phase",
        type_="Negative",
        priority="High",
        preconditions=f"{tgt} becomes unreachable after source extraction succeeds",
        steps=(
            "1. Trigger integration\n"
            "2. Verify data extracted from source successfully\n"
            "3. Load phase fails\n"
            "4. Verify extracted data is not lost"
        ),
        expected="Retry per policy; on failure, extracted data preserved in staging/queue; status = Error; notification sent",
    )

    add(
        scenario=f"Source authentication failure ({src_auth})",
        type_="Negative",
        priority="High",
        preconditions=f"Invalid or expired credentials configured for {src}",
        steps=(
            "1. Trigger integration\n"
            "2. Verify authentication fails\n"
            "3. Verify no data is pulled"
        ),
        expected="Clear auth failure message logged; integration stops; notification sent; no partial data loaded to target",
    )

    add(
        scenario=f"Target authentication failure ({tgt_auth})",
        type_="Negative",
        priority="High",
        preconditions=f"Invalid or expired credentials configured for {tgt}",
        steps=(
            "1. Trigger integration\n"
            "2. Source extraction succeeds\n"
            "3. Target load authentication fails"
        ),
        expected="Auth failure logged with clear message; extracted data preserved; status = Error",
    )

    # ═══ ERROR HANDLING & RECOVERY ═══
    add(
        scenario="Transient failure — retry mechanism succeeds on second attempt",
        type_="Error Handling",
        priority="High",
        preconditions="Simulate a transient failure (temporary network glitch, brief service outage)",
        steps=(
            "1. Trigger integration\n"
            "2. Inject transient failure on first attempt\n"
            "3. Verify retry occurs per configured policy\n"
            "4. Verify eventual success on retry"
        ),
        expected="Integration retries (typically 3 attempts with exponential backoff); succeeds on subsequent attempt; final status = Completed",
    )

    add(
        scenario="Notification sent on successful run",
        type_="Positive",
        priority="Medium",
        preconditions="Notification recipients configured; email/webhook delivery working",
        steps=(
            "1. Run integration successfully\n"
            "2. Check notification delivered to configured recipients"
        ),
        expected="Success notification received with run summary: record count, duration, run ID",
    )

    add(
        scenario="Notification sent on failure",
        type_="Error Handling",
        priority="High",
        preconditions="Notification recipients configured; failure scenario simulated",
        steps=(
            "1. Trigger integration\n"
            "2. Inject failure\n"
            "3. Check failure notification delivered"
        ),
        expected="Failure notification received with error details, run ID, timestamp, affected record count",
    )

    add(
        scenario="Run log entry created for every execution",
        type_="Positive",
        priority="Medium",
        preconditions="Process monitoring / run log table configured",
        steps=(
            "1. Trigger any run (success or failure)\n"
            "2. Query the run log table"
        ),
        expected="Entry present with: run_id, start_time, end_time, status, record_count, error_message (if any)",
        comments="Audit requirement",
    )

    # ═══ WORK-TYPE-SPECIFIC ═══
    if "file" in src_type:
        add(
            scenario="Corrupted or unreadable source file",
            type_="Negative",
            priority="High",
            preconditions=f"Malformed file placed in {src} pickup location",
            steps=(
                "1. Trigger integration\n"
                "2. Verify file fails validation\n"
                "3. File moved to error/quarantine folder"
            ),
            expected="Bad file rejected; integration continues with other files (if batch); error logged with filename",
            comments="File-source specific (T1 / T4)",
        )
        add(
            scenario="File encoding or minor format variance",
            type_="Boundary",
            priority="Medium",
            preconditions="File with unexpected encoding (e.g. UTF-8 with BOM) or minor format variance",
            steps=(
                "1. Place file in pickup location\n"
                "2. Trigger integration\n"
                "3. Verify handling"
            ),
            expected="Either processed successfully if within tolerance, or rejected with clear error",
        )

    if "api" in src_type:
        add(
            scenario="API pagination — full result set retrieved across multiple pages",
            type_="Positive",
            priority="High",
            preconditions=f"{src} returns paginated results totalling more than one page",
            steps=(
                "1. Trigger integration\n"
                "2. Verify all pages fetched (follow @odata.nextLink / cursor / offset)\n"
                "3. Compare total processed count to source total"
            ),
            expected="All pages consumed; total record count matches source total; no records missed at page boundaries",
            comments="API-source specific (T2)",
        )
        add(
            scenario="API rate limit exceeded (HTTP 429)",
            type_="Negative",
            priority="Medium",
            preconditions=f"{src} rate limit is exhausted mid-run",
            steps=(
                "1. Exhaust rate limit\n"
                "2. Verify integration handles 429 response\n"
                "3. Verify backoff behavior"
            ),
            expected="Integration honors Retry-After header; backs off; eventually completes without data loss",
        )
        add(
            scenario="OAuth token expires mid-run",
            type_="Error Handling",
            priority="High",
            preconditions="OAuth token near expiry; long-running extract",
            steps=(
                "1. Trigger integration\n"
                "2. Let token expire during a page fetch\n"
                "3. Verify token refresh is attempted"
            ),
            expected="Token refreshed automatically; run continues without failure",
        )

    if "api" in tgt_type:
        add(
            scenario=f"Target API returns 4xx client error",
            type_="Negative",
            priority="High",
            preconditions=f"Send malformed or invalid payload to {tgt} API",
            steps=(
                "1. Trigger with a bad payload\n"
                "2. Capture 4xx response\n"
                "3. Verify handling"
            ),
            expected="4xx errors are NOT retried (permanent errors); failed record moved to error queue; notification sent",
            comments="API-target specific (T5)",
        )
        add(
            scenario=f"Target API returns 5xx server error",
            type_="Negative",
            priority="High",
            preconditions=f"{tgt} API returns 5xx during load",
            steps=(
                "1. Trigger integration\n"
                "2. Simulate 5xx from target\n"
                "3. Verify retry behavior"
            ),
            expected="5xx errors trigger retry with exponential backoff; on persistent failure, status = Error",
        )

    if "database" in tgt_type:
        add(
            scenario=f"Target table constraint violation ({target_tbl})",
            type_="Negative",
            priority="High",
            preconditions=f"Record violates a constraint (FK, unique, not-null) in {target_tbl}",
            steps=(
                "1. Trigger integration\n"
                "2. Load fails on constraint violation"
            ),
            expected="Failed row identified with the constraint name; other rows continue processing; error logged with row identifier",
            comments="DB-target specific (T1 / T3)",
        )

    # ═══ VOLUME / PERFORMANCE ═══
    add(
        scenario="Large volume processing — performance under peak load",
        type_="Boundary",
        priority="Medium",
        preconditions=f"Source contains volume at or above expected peak: TBD",
        steps=(
            "1. Load source with peak-volume data\n"
            "2. Trigger integration\n"
            "3. Measure end-to-end duration\n"
            f"4. Verify SLA compliance: {sla}"
        ),
        expected=f"Integration completes within SLA ({sla}); no memory or timeout errors; all records processed",
        comments="Performance test",
    )

    return cases


def generate_test_cases_xlsx(canonical):
    """Return a BytesIO of a formatted test-cases workbook (.xlsx)."""
    project_name = _v(canonical.get("project", {}).get("name"), "Integration")
    wt_label     = _v(canonical.get("integration", {}).get("workType"))
    src          = _v(canonical.get("source", {}).get("system"))
    tgt          = _v(canonical.get("target", {}).get("system"))

    cases = _build_test_cases(canonical)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Row 1 — title (merged across all columns)
    title_cell = ws.cell(row=1, column=1, value=f"Test Cases — {project_name}")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 26

    # Row 2 — metadata line (merged)
    meta = f"Work Type: {wt_label}   |   Source: {src}   |   Target: {tgt}   |   Generated: {datetime.now().strftime('%Y-%m-%d')}"
    meta_cell = ws.cell(row=2, column=1, value=meta)
    meta_cell.font = META_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    # Row 4 — column headers
    HEADER_ROW = 4
    for col_idx, (name, width, _align) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 26

    # Data rows
    row = HEADER_ROW + 1
    for i, tc in enumerate(cases):
        values = [
            tc["id"], tc["scenario"], tc["type"], tc["priority"],
            tc["preconditions"], tc["steps"], tc["expected"],
            "", "", tc.get("comments", ""),
        ]
        alt = (i % 2 == 1)
        for col_idx, (val, (_, _, align)) in enumerate(zip(values, COLUMNS), start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                wrap_text=True,
                vertical='top',
                horizontal='center' if align == 'center' else 'left',
            )
            # TBD highlighting wins over alt-row shading
            if _has_tbd(val):
                cell.fill = TBD_FILL
                cell.font = TBD_FONT
            else:
                cell.font = CELL_FONT
                if alt:
                    cell.fill = ALT_FILL
        row += 1

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = f'A{HEADER_ROW + 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
